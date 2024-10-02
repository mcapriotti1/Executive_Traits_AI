import openpyxl
from openpyxl import load_workbook
import os
import requests
import pytesseract
from PIL import Image
import tiktoken
import asyncio
import nest_asyncio
from pyppeteer import connect
import pyperclip
import pandas as pd
import re
from pyppeteer.errors import ElementHandleError
import shutil

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''Extracting Data'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def extract_columns_to_file(excel_file_name, max_rows=100, workbook_folder="ExcelSpreadsheets", output_folder="excel_data", output_file_name="link_data.txt"):
    # Construct the full path to the Excel file
    workbook_path = os.path.join(workbook_folder, excel_file_name)

    # Load the workbook and select the active sheet
    book = load_workbook(workbook_path)
    sheet = book.active

    # Convert column letters to indexes (1-based)
    start_col_index = openpyxl.utils.column_index_from_string('H')
    end_col_index = openpyxl.utils.column_index_from_string('Z')
    check_col_index = openpyxl.utils.column_index_from_string('AB')

    # Initialize a list to collect rows of data
    rows_data = []

    # Flag to indicate the start of data collection
    start_collecting = False
    rows_collected = 0

    # Variable to store the starting row
    start_row = None

    # Iterate over the rows
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract the values from the columns I through V
        columns_I_to_V = row[start_col_index-1:end_col_index]
        # Extract the value from column W
        column_W = row[check_col_index-1]

        # Check if columns I through V are all empty and column W has data
        if all(cell is None for cell in columns_I_to_V) and column_W is not None:
            if not start_collecting:
                start_row = row_index  # Record the starting row
                start_collecting = True

        # If data collection has started, collect the data
        if start_collecting and rows_collected < max_rows:
            # Stop collecting if column W is empty
            if column_W is None:
                break

            # Start with the value in column W
            row_data = [str(column_W) if column_W is not None else '']

            # Check up to 2 columns to the right of W (X, Y)
            for offset in range(1, 3):
                col_index = check_col_index - 1 + offset
                if col_index < len(row) and row[col_index] is not None:
                    row_data.append(str(row[col_index]))
                else:
                    row_data.append('')

            rows_data.append(' '.join(row_data).strip())
            rows_collected += 1

        # Stop collecting after max_rows rows
        if rows_collected >= max_rows:
            break

    # Create the output directory if it does not exist
    os.makedirs(output_folder, exist_ok=True)

    # Define the output file path
    output_file_path = os.path.join(output_folder, output_file_name)

    # Check if there is any data in the rows
    if rows_data:
        # Open a text file to write the rows data
        with open(output_file_path, "w") as file:
            for row in rows_data:
                file.write(row + "\n")
        return len(rows_data), start_row
    else:
        print("No data found to write.")
        return 0, None

def process_excel_and_extract_columns(start_row, excel_file_name):
    # Predefined folder and output directory within the function
    excel_folder = "ExcelSpreadsheets"
    output_dir = "excel_data"
    link_data_file = "link_data.txt"
    
    # Function to extract a single column from the workbook and write to a file
    def extract_column_to_file(workbook_path, column_letter, output_file_path, num_rows=None):
        # Load the workbook and select the active sheet
        book = load_workbook(workbook_path)
        sheet = book.active

        # Initialize a list to collect column data
        column_data = []

        # Convert start_row to 0-based index for Python list slicing
        start_index = start_row - 1
        end_index = (start_row + num_rows - 1) if num_rows else None

        # Collect the column data, starting from start_row
        for cell in sheet[column_letter][start_index:end_index]:
            column_data.append(cell.value)

        # Check if there is any data in the column
        if column_data:
            # Write the column data to the text file
            with open(output_file_path, "w") as file:
                for value in column_data:
                    file.write("" if value is None else str(value))
                    file.write("\n")
        else:
            print(f"No data found in column {column_letter}")

    # Construct the full path to the Excel file
    workbook_path = os.path.join(excel_folder, excel_file_name)

    # Create the directory if it does not exist
    os.makedirs(output_dir, exist_ok=True)

    # Define the output file paths for the columns
    names_data_file = os.path.join(output_dir, "names_data.txt")
    role_data_file = os.path.join(output_dir, "role_data.txt")
    company_data_file = os.path.join(output_dir, "company_data.txt")

    # Number of rows to extract (obtained from link_data extraction previously)
    num_rows = None
    try:
        with open(os.path.join(output_dir, link_data_file), "r") as file:
            num_rows = sum(1 for _ in file)
    except FileNotFoundError:
        print(f"{link_data_file} file not found. Make sure to extract link_data first.")
        return

    # Extract data from the specific columns if num_rows is defined
    if num_rows:
        extract_column_to_file(workbook_path, 'A', names_data_file, num_rows)
        extract_column_to_file(workbook_path, 'D', role_data_file, num_rows)
        extract_column_to_file(workbook_path, 'E', company_data_file, num_rows)


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''Image To Text'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def download_images(urls_file, output_dir):
    def download_image(image_url, file_dir):
        response = requests.get(image_url)
        if response.status_code == 200:
            directory = os.path.dirname(file_dir)
            if not os.path.exists(directory):
                os.makedirs(directory)
            with open(file_dir, "wb") as fp:
                fp.write(response.content)

    def download_images_from_file(file_with_urls, download_dir):
        with open(file_with_urls, 'r') as file:
            urls = file.readlines()

        for index, line in enumerate(urls):
            urls_in_line = line.strip().split()  # Split line into URLs by spaces
            for sub_index, url in enumerate(urls_in_line):
                if url:
                    if len(urls_in_line) > 1:
                        file_name = f"image_{index + 1}.{sub_index + 1}.jpg"
                    else:
                        file_name = f"image_{index + 1}.jpg"
                    file_path = os.path.join(download_dir, file_name)
                    download_image(url, file_path)

    download_images_from_file(urls_file, output_dir)

def convert_images_to_text_files(image_dir, output_dir):
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    for image_file in os.listdir(image_dir):
        if image_file.endswith(".jpg"):
            image_file_path = os.path.join(image_dir, image_file)
            text_file_name = os.path.splitext(image_file)[0] + ".txt"
            text_file_path = os.path.join(output_dir, text_file_name)
            
            if os.path.exists(image_file_path):
                image = Image.open(image_file_path)
                raw_text = str(pytesseract.image_to_string(image))
                
                with open(text_file_path, "w") as text_file:
                    text_file.write(raw_text)
            else:
                pass  # No action needed for missing files

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''File Formatting'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def combine_files(folder, file1, file2, file3, output_file):
    # Paths to the input files
    file1_path = os.path.join(folder, file1)
    file2_path = os.path.join(folder, file2)
    file3_path = os.path.join(folder, file3)

    # Open the output file for writing with UTF-8 encoding
    with open(output_file, 'w', encoding='utf-8') as outfile:
        with open(file1_path, 'r', encoding='utf-8') as f1, open(file2_path, 'r', encoding='utf-8') as f2, open(file3_path, 'r', encoding='utf-8') as f3:
            # Read lines from each file
            lines1 = f1.readlines()
            lines2 = f2.readlines()
            lines3 = f3.readlines()
            
            # Determine the maximum number of lines
            max_lines = max(len(lines1), len(lines2), len(lines3))
            
            # Iterate through each line and combine them
            for i in range(max_lines):
                line1 = lines1[i].strip() if i < len(lines1) else ''
                line2 = lines2[i].strip() if i < len(lines2) else ''
                line3 = lines3[i].strip() if i < len(lines3) else ''
                # Combine lines with a '|' and ensure there's a space between them
                combined_line = f"Ҩ{line1 or ' '}Ҩ{line2 or ' '}Ҩ{line3 or ' '}Ҩ\n"
                outfile.write(combined_line)

# Folder containing the input files
folder = "excel_data"

# Names of the input files
file1 = "company_data.txt"
file2 = "names_data.txt"
file3 = "role_data.txt"

# Path to the output file
output_file = os.path.join(folder, "combined_excel_data.txt")

def combine_text_files(text_files_folder, output_folder):
    """Combine text from image files and save them in separate output files."""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Dictionary to store combined text for each base image
    combined_texts = {}

    # Iterate over files in the text files folder
    for text_file in os.listdir(text_files_folder):
        if text_file.endswith(".txt"):
            # Extract base name (e.g., image_1)
            base_name = text_file.split('.')[0]  # E.g., image_1.1 -> image_1
            if base_name not in combined_texts:
                combined_texts[base_name] = []
            
            # Read the text file and append content to the list
            file_path = os.path.join(text_files_folder, text_file)
            with open(file_path, 'r') as file:
                content = file.read().strip()
                combined_texts[base_name].append(content)
    
    # Write combined text to new files and print results
    for base_name, texts in combined_texts.items():
        combined_content = '\n'.join(texts)  # Combine all texts, separated by new lines
        output_file_name = f"combined_{base_name}.txt"  # E.g., combined_image_1.txt
        output_file_path = os.path.join(output_folder, output_file_name)
        with open(output_file_path, 'w') as file:
            file.write(combined_content)
        #print(f"Contents of {output_file_name}:\n{combined_content}\n")

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''GPT Formatting'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def count_tokens(text):
    enc = tiktoken.get_encoding("cl100k_base")
    tokens = enc.encode(text)
    return len(tokens)

def read_file_with_encoding(file_path):
    """Attempt to read a file with UTF-8 encoding first, fall back to latin-1 if it fails."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.readlines()
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='latin-1') as file:
            return file.readlines()

def combine_all(excel_data_file, images_folder, custom_message, output_folder = 'gpt_input'):
    """Combine each line from combined_excel_data.txt with corresponding combined_image_(number).txt and save the combined files."""

    # Ensure the output directory exists
    os.makedirs(output_folder, exist_ok=True)

    # Read lines from combined_excel_data.txt
    excel_file_path = os.path.join('excel_data', excel_data_file)
    excel_lines = read_file_with_encoding(excel_file_path)
    
    # Dictionary to store image texts
    image_texts = {}

    # Iterate over files in the text files folder
    for image_file in os.listdir(images_folder):
        if image_file.startswith("combined_image_") and image_file.endswith(".txt"):
            number = image_file.split('_')[2].split('.')[0]  # Extract the number
            image_file_path = os.path.join(images_folder, image_file)
            image_texts[number] = read_file_with_encoding(image_file_path)
    
    combined_text = ""
    current_token_count = count_tokens(custom_message)  # Include custom message in token count
    file_number = 1
    global_counter = 1  # Start the global counter at 1

    for index, line in enumerate(excel_lines):
        image_number = str(index + 1)  # Image numbers are 1-based
        image_text = ' '.join(image_texts.get(image_number, [''])).strip()  # Get corresponding image text or empty string
        combined_content = f"{global_counter}Ɬ{line.strip()} {image_text} Ɬ\n"
        global_counter += 1  # Increment the global counter after each item
        combined_content_tokens = count_tokens(combined_content)

        if current_token_count + combined_content_tokens <= 4000:
            combined_text += combined_content
            current_token_count += combined_content_tokens
        else:
            # Save the current combined text to a file with the custom message at the beginning
            combined_file_name = f"gpt_combined_input_{file_number}.txt"
            combined_file_path = os.path.join(output_folder, combined_file_name)
            with open(combined_file_path, 'w', encoding='utf-8') as file:
                file.write(custom_message + combined_text.strip())
            # Print the combined text content
            #print(f"Contents of {combined_file_name}:\n{combined_text.strip()}\n")

            # Reset for the next file
            combined_text = combined_content
            current_token_count = combined_content_tokens + count_tokens(custom_message)
            file_number += 1

    # Save any remaining combined text to a final file with the custom message at the beginning
    if combined_text:
        combined_file_name = f"gpt_combined_input_{file_number}.txt"
        combined_file_path = os.path.join(output_folder, combined_file_name)
        with open(combined_file_path, 'w', encoding='utf-8') as file:
            file.write(custom_message + combined_text.strip())
        # Print the combined text content
        # print(f"Contents of {combined_file_name}:\n{combined_text.strip()}\n")


excel_data_file = "combined_excel_data.txt"
images_folder = r'text_files'  # Folder containing combined_image_(number).txt files
custom_message = r"I'm going to give you information and I want you to read it and answer questions. The text was transferred from image-text so some of it might not make much sense, use your best judgment to guess what certain words are. I will give you preliminary information in this format: (number)ҨcompanyҨnameҨpositionҨ. The number represnts the assinged number for each person. This will be followed by text that I want you to use to answer the following questions. I will give you multiple different texts separated by this symbol Ɬ, answer the questions for each text individually. You MUST BE 80% confident to answer, otherwise leave it blank. 1. Description: The company mentioned in prelimnary information also contains the location, just look at the first part of the text which doesnt have numbers as that contains the company name. If the company in preliminary information is present at all, check all text to ensure weather it is presnt or not present in the text, or a only certain parts of a company name, or a company which seems to have a similar name, say yes (sometimes it will be abberviated). Most of the time this answer will be yes. Answer: Yes or No 2. Co: Description: Do nothing with this. Leave blank. 3. Description: Based on their line of work, is it the same industry, a different industry, or multiple industries compared to the earlier company? Use the name of the listed company and see if it mentions that they did that work. Most of the time this answer will be same. Answer: Same, Different, or Multiple 4. At Time of Death - Still at Firm: Description: Were they still at the firm when they died, or had they already retired/changed to a different company? Most of the time this answer will be yes, only answer no if it specifically says they left the company before their death or retired. Answer: Yes or No 5. Title Change: Description: Is their title at the company different from the one listed in the preliminary information? If the answer to this is no leave it blank. Answer: Yes or leave blank 6. New Title: Description: Only answer if the Title Change answer was Yes. What is their new title? Answer: New Title 7. Year Started at Firm, or Years at Firm: Description: What year did they first start working at the firm listed, or how many years have they worked there? If this is not provided leave it blank. Answer: Year or Years 8. Year Started in Field, or Years in Field: Description: What year did they first start working in their field/industry, or how many years have they worked in their field? If this is not provided leave it blank. Answer: Year or Years 9. Initial Title: Description: What was their initial title when they first joined the company? Answer: Initial Title 10. Director's Relation to Our Firm or Sector: Description: All the following must refer to the company given in the preliminary information or the same sector as the company in the prelimnary information, otherwise leave blank. Work their way up to an executive position over time (Career), founded the company (Founder), family member who founded the company (Family Founder), serve as the lawyer for the company (Lawyer), travel around in meetings for the company (Itinerant), or previously founded/worked at a company and then instantly became an executive at the preliminary company (Capitalist)? Answer: Career, Founder, Family Founder, Itinerant, Capitalist, or Lawyer 11. Additional Info Related to Firm: Description: Did they have a bank seat at their firm? If No just leave it blank. Answer: Yes or leave blank 12. Other Directorships Mentioned: Description: Were there other companies they were associated with? Answer: Yes or No 13. Director's Main Sector: Description: What was the sector the director was primarily associated with? Answer: Sector 14. If Banker - Cbank or Ibank: Description: If their sector was banking, were they more associated with commercial or investment banking? Answer: Cbank or Ibank 15. University/College: Description: If they attended a university or college, what was the name of the school? Answer: Name of the School 16. Director's Main Company: Description: Is the company listed in the preliminary information their main company? Answer: Yes or No 17. Other Affiliated Companies: Description: If the answer to the previous question was No, name the company the director seemed most associated with. Answer: Company Name 18. Former CEO Now Director: Description: Were they formerly a CEO and are now just a director? If the answer is no just leave blank. Answer: Yes or leave blank 19. Additional Notes: Description: Leave blank most of the time, unless something goes very wrong, for example if there is almost no information, write 'no information' Answer: 'no information'. Instructions for Output Formatting: Answer all questions in the specified format. Do not repeat the titles of the answer choices in your response, or any text besides the exact output format shown later. Do not add extra text, line spaces, or enters between different sets of answers. The answers should be a continuous block of text, not a list format. For each question, use the question number followed by a period, then provide the answer directly after the period. If there is no answer for a question, leave it blank but still include the question number followed by a period. Always number from 1 to 19, ensuring that each question is accounted for. Include the assigned number in the format Ɬ(number)Ɬ at the beginning of each block of answers. Do not include any titles, like Text_1. Follow this Output exactly: Ɬ(assigned number)Ɬ1. answer 2. answer 3. answer 4. answer 5. answer 6. answer 7. answer 8. answer 9. answer 10. answer 11. answer 12. answer 13. answer 14. answer 15. answer 16. answer 17. answer 18. answer 19. answer"

def process_files(input_folder, output_folder):
    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Process each file in the input folder
    for filename in os.listdir(input_folder):
        if filename.startswith('gpt_combined_input_') and filename.endswith('.txt'):
            # Read the content of the input file
            with open(os.path.join(input_folder, filename), 'r', encoding='utf-8') as file:
                content = file.read()

            # Replace '@' with 'a'
            modified_content = content.replace('@', 'a')

            # Construct the output filename
            output_filename = filename.replace('gpt_combined_input_', 'gpt_removed_')

            # Write the modified content to the output file
            with open(os.path.join(output_folder, output_filename), 'w', encoding='utf-8') as file:
                file.write(modified_content)

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''Running Through GPT'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

nest_asyncio.apply()

async def query_gpt(input_text, page):
    try:
        await asyncio.sleep(2)
        
        # Copy the input text to clipboard
        pyperclip.copy(input_text)
        
        # Wait for the GPT text box to be available using the provided selector
        await page.waitForSelector('#prompt-textarea > p')
        
        # Ensure the text box is focused
        await page.focus('#prompt-textarea > p')
        
        # Use JavaScript to directly set the text box value
        await page.evaluate('''(text) => {
            const textarea = document.querySelector('#prompt-textarea > p');
            if (textarea) {
                textarea.innerText = text;
            }
        }''', input_text)
        
        # Add a short delay before pressing Enter
        await asyncio.sleep(1)

        # Make sure Shift is not held down
        await page.keyboard.up('Shift')

        # Press Enter to submit
        await page.keyboard.press('Enter')

        # Wait for the response to be generated
        await asyncio.sleep(30)

        return "Response processed"

    except Exception as e:
        print(f"An error occurred: {e}")
        return ""

async def process_gpt_files(gpt_input_folder='gpt_final_input'):
    # Get list of input files
    input_files = [f for f in os.listdir(gpt_input_folder) if f.startswith('gpt_removed_') and f.endswith('.txt')]
    
    # Sort the files in numerical ascending order
    input_files.sort(key=lambda x: int(x.split('_')[-1].split('.')[0]))
    
    try:
        browser = await connect({'browserURL': 'http://localhost:9222'})
        pages = await browser.pages()
        page = pages[0]  # Assume the target page is the first one
        
        for input_file in input_files:
            input_file_path = os.path.join(gpt_input_folder, input_file)
            
            try:
                with open(input_file_path, 'r', encoding='utf-8') as file:
                    input_text = ' '.join(line.strip() for line in file.readlines())
            except Exception as e:
                print(f"Error reading file {input_file_path}: {e}")
                continue
            
            # Send the input text to ChatGPT
            response = await query_gpt(input_text, page)
            
        await browser.disconnect()  # Properly disconnect the browser
    
    except Exception as e:
        print(f'Error occurred: {e}')

# Allow asyncio to run in Jupyter or similar environments
nest_asyncio.apply()

# Create the output directory if it doesn't exist
output_dir = 'gpt_output'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

async def scrape_text_from_website():
    # Connect to the existing browser instance
    browser = await connect(browserURL='http://localhost:9222')
    page = await browser.pages()  # Get all open pages
    current_page = page[0]  # Assume the first page is the one we want to scrape

    output_file_path = os.path.join(output_dir, 'scraped_text.txt')

    # Open the file in write mode (or append if you prefer)
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        n = 2  # Start from article[2]
        while True:
            try:
                # Construct the XPath for the current article
                article_xpath = f'/html/body/div[1]/div[2]/main/div[1]/div[1]/div/div/div/div/article[{n}]'
                p_index = 1  # Start checking for <p> from index 1
                
                # Check if the article exists
                article_exists = await current_page.xpath(article_xpath)
                if not article_exists:
                    break  # Exit the loop if no more articles

                # Flag to indicate if we found any paragraphs
                paragraphs_found = False

                # Loop to check <p> tags within the current article
                while True:
                    try:
                        p_xpath = f'{article_xpath}/div/div/div[2]/div/div[1]/div/div/div/p[{p_index}]'
                        p_elements = await current_page.xpath(p_xpath)
                        
                        if p_elements:
                            # If the <p> element exists, extract and save its text
                            p_text = await current_page.evaluate('(element) => element.textContent', p_elements[0])
                            output_file.write(p_text + '\n')  # Write the paragraph to the file
                            paragraphs_found = True  # Mark that we found paragraphs
                            p_index += 1  # Check for the next <p> index
                        else:
                            break  # Exit if no more <p> elements are found at this index

                    except ElementHandleError:
                        break  # Exit if there is an error accessing the <p> element

                # If no specific paragraphs were found, check for the default <p>
                if not paragraphs_found:
                    p_default_xpath = f'{article_xpath}/div/div/div[2]/div/div[1]/div/div/div/p'
                    p_default_elements = await current_page.xpath(p_default_xpath)
                    
                    if p_default_elements:
                        p_default_text = await current_page.evaluate('(element) => element.textContent', p_default_elements[0])
                        output_file.write(p_default_text + '\n')

                n += 2  # Increment to the next article
            except Exception as e:
                output_file.write(f'An error occurred: {e}\n')
                break
                
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''Folder Formatting'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def move_folders(folders_to_move, target_folder):
    """
    Move specified folders to a target folder.

    Parameters:
    folders_to_move (list): List of folder paths to move.
    target_folder (str): Path to the target folder where the folders will be moved.
    """
    # Create the target folder if it doesn't exist
    os.makedirs(target_folder, exist_ok=True)

    # Move each specified folder to the target folder
    for folder in folders_to_move:
        try:
            # Get the folder name
            folder_name = os.path.basename(folder)
            
            # Define the destination path
            destination = os.path.join(target_folder, folder_name)
            
            # Move the folder
            shutil.move(folder, destination)
        except Exception as e:
            # You can log the error instead of printing it
            # For example, you could use logging
            pass  # Optionally handle errors here (e.g., logging)

# Example usage:
if __name__ == "__main__":
    folders = [
        "text_files",
        "gpt_input",
        "gpt_final_input",
        "ExcelSpreadsheets",
        "excel_data",
        "downloaded_images",
    ]
    target = "All_Inputs"


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''Delete All Files'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def delete_all_files_in_folders(folder_paths):
    """
    Delete all files and subfolders in the specified folders, and then delete the folders themselves.
    
    :param folder_paths: List of paths to the folders to be deleted.
    """
    for folder_path in folder_paths:
        if not os.path.exists(folder_path):
            continue  # Skip if folder does not exist
        
        try:
            # Remove all contents (files and subfolders)
            shutil.rmtree(folder_path)
        except Exception as e:
            # Handle the exception as needed (e.g., log it or raise)
            pass  # Or log the error without printing

# Example usage
if __name__ == "__main__":
    folders_to_delete = [
        "text_files",
        "gpt_input",
        "gpt_final_input",
        "ExcelSpreadsheets",
        "excel_data",
        "downloaded_images",
    ]